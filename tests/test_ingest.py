"""The ingestion layer: format conversion, sanitization, and cache invalidation."""

import json
import os
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

from insight_agent.ingest import TableSource, discover_tables, scan_tables


def _names(sources: list[TableSource]) -> set[str]:
    return {source.table_name for source in sources}


def test_csv_file_loads_directly(tmp_path: Path) -> None:
    (tmp_path / "orders.csv").write_text("a,b\n1,2\n", encoding="utf-8")
    sources = discover_tables(tmp_path)
    assert _names(sources) == {"orders"}
    assert sources[0].csv_path == tmp_path / "orders.csv"


def test_xlsx_with_two_sheets_becomes_two_tables(tmp_path: Path) -> None:
    workbook = Workbook()
    first = workbook.active
    first.title = "Sheet1"
    first.append(["a", "b"])
    first.append([1, 2])
    second = workbook.create_sheet("Sheet2")
    second.append(["c", "d"])
    second.append([3, 4])
    workbook.save(tmp_path / "book.xlsx")

    sources = discover_tables(tmp_path)
    assert _names(sources) == {"book_sheet1", "book_sheet2"}
    for source in sources:
        assert source.csv_path.exists()
        assert source.csv_path.parent == tmp_path / ".converted"


def test_xlsx_with_single_sheet_uses_stem(tmp_path: Path) -> None:
    workbook = Workbook()
    workbook.active.append(["a", "b"])
    workbook.active.append([1, 2])
    workbook.save(tmp_path / "single.xlsx")

    sources = discover_tables(tmp_path)
    assert _names(sources) == {"single"}


def test_txt_with_semicolon_delimiter_is_sniffed(tmp_path: Path) -> None:
    (tmp_path / "data.txt").write_text("a;b;c\n1;2;3\n4;5;6\n", encoding="utf-8")
    sources = discover_tables(tmp_path)
    assert _names(sources) == {"data"}
    frame = pd.read_csv(sources[0].csv_path)
    assert list(frame.columns) == ["a", "b", "c"]
    assert frame.shape == (2, 3)


def test_tsv_with_tab_delimiter_is_sniffed(tmp_path: Path) -> None:
    (tmp_path / "data.tsv").write_text("a\tb\tc\n1\t2\t3\n4\t5\t6\n", encoding="utf-8")
    sources = discover_tables(tmp_path)
    assert _names(sources) == {"data"}
    frame = pd.read_csv(sources[0].csv_path)
    assert list(frame.columns) == ["a", "b", "c"]
    assert frame.shape == (2, 3)


def test_json_round_trips(tmp_path: Path) -> None:
    records = [{"a": 1, "b": "x"}, {"a": 2, "b": "y"}]
    (tmp_path / "records.json").write_text(json.dumps(records), encoding="utf-8")
    sources = discover_tables(tmp_path)
    assert _names(sources) == {"records"}
    frame = pd.read_csv(sources[0].csv_path)
    assert list(frame["a"]) == [1, 2]
    assert list(frame["b"]) == ["x", "y"]


def test_parquet_round_trips(tmp_path: Path) -> None:
    frame = pd.DataFrame({"a": [1, 2], "b": ["x", "y"]})
    frame.to_parquet(tmp_path / "table.parquet")
    sources = discover_tables(tmp_path)
    assert _names(sources) == {"table"}
    loaded = pd.read_csv(sources[0].csv_path)
    assert list(loaded["a"]) == [1, 2]
    assert list(loaded["b"]) == ["x", "y"]


def test_name_sanitization_handles_spaces_dashes_and_leading_digit(tmp_path: Path) -> None:
    (tmp_path / "My Sales-Data.csv").write_text("a,b\n1,2\n", encoding="utf-8")
    (tmp_path / "2024_report.csv").write_text("a,b\n1,2\n", encoding="utf-8")
    sources = discover_tables(tmp_path)
    assert _names(sources) == {"my_sales_data", "t_2024_report"}


def test_collision_suffixes_assigned_by_sorted_file_order(tmp_path: Path) -> None:
    (tmp_path / "a.csv").write_text("x\n1\n", encoding="utf-8")
    (tmp_path / "a.txt").write_text("x\n2\n", encoding="utf-8")
    sources = discover_tables(tmp_path)
    by_name = {source.table_name: source for source in sources}
    assert set(by_name) == {"a", "a_2"}
    assert by_name["a"].csv_path == tmp_path / "a.csv"
    assert by_name["a_2"].csv_path.name.startswith("a")


def test_unsupported_extension_is_skipped_without_error(tmp_path: Path) -> None:
    (tmp_path / "notes.md").write_text("# hello", encoding="utf-8")
    (tmp_path / "orders.csv").write_text("a,b\n1,2\n", encoding="utf-8")
    sources = discover_tables(tmp_path)
    assert _names(sources) == {"orders"}


def test_converted_directory_contents_are_never_scanned_as_sources(tmp_path: Path) -> None:
    (tmp_path / "orders.txt").write_text("a,b\n1,2\n", encoding="utf-8")
    discover_tables(tmp_path)
    converted_dir = tmp_path / ".converted"
    assert converted_dir.exists()
    # A second scan must not pick up the cached CSV inside .converted as its
    # own source table.
    sources = discover_tables(tmp_path)
    assert _names(sources) == {"orders"}


def test_cache_is_reconverted_when_source_is_newer(tmp_path: Path) -> None:
    source = tmp_path / "data.txt"
    source.write_text("a,b\n1,2\n", encoding="utf-8")
    sources = discover_tables(tmp_path)
    cache_path = sources[0].csv_path
    original_mtime = cache_path.stat().st_mtime

    source.write_text("a,b\n9,9\n", encoding="utf-8")
    newer = original_mtime + 10
    os.utime(source, (newer, newer))

    sources_again = discover_tables(tmp_path)
    frame = pd.read_csv(sources_again[0].csv_path)
    assert list(frame["a"]) == [9]


def test_cache_is_reused_when_source_is_not_newer(tmp_path: Path) -> None:
    source = tmp_path / "data.txt"
    source.write_text("a,b\n1,2\n", encoding="utf-8")
    sources = discover_tables(tmp_path)
    cache_path = sources[0].csv_path
    cache_mtime_before = cache_path.stat().st_mtime

    sources_again = discover_tables(tmp_path)
    assert sources_again[0].csv_path.stat().st_mtime == pytest.approx(cache_mtime_before)
    frame = pd.read_csv(sources_again[0].csv_path)
    assert list(frame["a"]) == [1]


def test_missing_directory_returns_no_tables(tmp_path: Path) -> None:
    assert discover_tables(tmp_path / "does_not_exist") == []


def test_scan_reports_unsupported_and_failed_files_as_skipped(tmp_path: Path) -> None:
    (tmp_path / "orders.csv").write_text("a,b\n1,2\n", encoding="utf-8")
    (tmp_path / "run.bat").write_text("echo hello\n", encoding="utf-8")
    (tmp_path / "broken.xlsx").write_text("not a real workbook", encoding="utf-8")

    discovery = scan_tables(tmp_path)

    assert _names(discovery.tables) == {"orders"}
    skipped = {entry.file: entry.reason for entry in discovery.skipped}
    assert set(skipped) == {"run.bat", "broken.xlsx"}
    assert "unsupported file type" in skipped["run.bat"]
    assert "conversion failed" in skipped["broken.xlsx"]
